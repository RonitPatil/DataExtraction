from langchain_astradb import AstraDBVectorStore
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.llms import HuggingFacePipeline
from langchain.schema import Document
from tenacity import retry, wait_random_exponential, stop_after_attempt
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.cell import MergedCell
import os
from local_models import get_local_embeddings, get_local_llm

def get_astra_vectorstore():
    embeddings = get_local_embeddings()
    if not embeddings:
        st.error("Failed to load embedding model")
        return None
    
    vectorstore = AstraDBVectorStore(
        embedding=embeddings,
        collection_name=os.getenv("ASTRA_DB_COLLECTION"),
        api_endpoint=os.getenv("ASTRA_DB_API_ENDPOINT"),
        token=os.getenv("ASTRA_DB_APPLICATION_TOKEN"),
    )
    return vectorstore

def get_rag_components():
    vectorstore = get_astra_vectorstore()
    if not vectorstore:
        return None, None
    
    llm = get_local_llm()
    if not llm:
        st.error("Failed to load LLM model")
        return None, None
    
    retriever = vectorstore.as_retriever()
    return llm, retriever

@retry(wait=wait_random_exponential(min=1, max=60), stop=stop_after_attempt(6))
def get_llm_response(llm, prompt):
    return llm.predict(prompt)

def safe_set_cell_value(ws, row, col, value):
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_cell.value = value
                    break
        else:
            cell.value = value
    except Exception as e:
        st.warning(f"Could not set value for cell ({row}, {col}): {e}")

def fill_excel_with_rag(excel_path, pdf_filename):
    llm, retriever = get_rag_components()
    if not llm or not retriever:
        st.error("Failed to initialize RAG components")
        return excel_path
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    df = pd.read_excel(excel_path)
    queries = df["Item Description"].astype(str).tolist()
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, query in enumerate(queries):
        progress = (idx + 1) / len(queries)
        progress_bar.progress(progress)
        status_text.text(f"Processing item {idx + 1}/{len(queries)}: {query[:50]}...")
        
        docs = retriever.get_relevant_documents(query)
        context_texts = [doc.page_content for doc in docs]
        
        pages = set()
        doc_names = set()
        
        for doc in docs:
            if hasattr(doc, 'metadata') and doc.metadata:
                if "page" in doc.metadata:
                    pages.add(doc.metadata["page"])
                elif "metadata" in doc.metadata and "page" in doc.metadata["metadata"]:
                    pages.add(doc.metadata["metadata"]["page"])
                
                if "document_name" in doc.metadata:
                    doc_names.add(doc.metadata["document_name"])
                elif "metadata" in doc.metadata and "document_name" in doc.metadata["metadata"]:
                    doc_names.add(doc.metadata["metadata"]["document_name"])
        
        pages = sorted(list(pages)) if pages else []
        doc_name = list(doc_names)[0] if doc_names else pdf_filename
        
        prompt = f"""
Given the following context extracted from a tender PDF, extract a consice 
description for the item only from the context below, DO NOT explain the item: "{query}".

Return only the relevant technical and commercial information related to the item.
Do not explain the item
If no information relevant to the item is found say "No information found in context"

Context:
{'-'*40}
{''.join(context_texts)}
{'-'*40}
"""
        details = get_llm_response(llm, prompt)
        
        safe_set_cell_value(ws, idx+2, 2, doc_name)
        safe_set_cell_value(ws, idx+2, 3, ", ".join(map(str, pages)))
        safe_set_cell_value(ws, idx+2, 5, details)
    
    progress_bar.empty()
    status_text.empty()
    
    out_path = excel_path.replace(".xlsx", "_filled.xlsx")
    wb.save(out_path)
    return out_path 